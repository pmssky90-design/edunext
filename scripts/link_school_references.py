from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from html import escape
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "output"
SCHOOL_MAP = ROOT / "data" / "school_region_map.json"
HOMEPAGE_MAP = ROOT / "data" / "school_official_homepages.json"


def valid_homepage(value: object) -> str:
    url = str(value or "").strip()
    parsed = urlparse(url)
    return url if parsed.scheme in {"http", "https"} and bool(parsed.netloc) and "홈페이지" not in parsed.netloc else ""


def read_homepages(xlsx_path: Path, wanted: set[tuple[str, str]]) -> dict[tuple[str, str], str]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(value or "").replace("\n", "").strip() for value in next(sheet.iter_rows(min_row=17, max_row=17, values_only=True))]
    indexes = {name: headers.index(name) for name in ("학교명", "주소", "홈페이지")}
    found: dict[tuple[str, str], str] = {}
    for row in sheet.iter_rows(min_row=18, values_only=True):
        school_name = str(row[indexes["학교명"]] or "").strip()
        address = str(row[indexes["주소"]] or "").strip()
        homepage = valid_homepage(row[indexes["홈페이지"]])
        if not school_name or not homepage:
            continue
        for key in wanted:
            official_name, city = key
            if school_name == official_name and city in address:
                if key in found and found[key] != homepage:
                    raise RuntimeError(f"conflicting homepage for {official_name} ({city})")
                found[key] = homepage
    workbook.close()
    return found


def source_link(href: str, label: str, external: bool) -> str:
    attrs = ' target="_blank" rel="noopener noreferrer external"' if external else ""
    arrow = "↗" if external else "→"
    return f'<a class="source-link" href="{escape(href, quote=True)}"{attrs}>{escape(label)} <span aria-hidden="true">{arrow}</span></a>'


def update_reference_section(html: str, link_html: str, label: str, replace_official: bool) -> str:
    official = re.compile(r'<li>학교 공식 홈페이지 주소(?:\(교육통계 수록\))?:\s*.*?</li>', re.I | re.S)
    replacement = f"<li>{label}: {link_html}</li>"
    if official.search(html):
        return official.sub(replacement, html, count=1)
    section = re.compile(r'(<h3\b[^>]*>자료 확인 기준</h3>\s*<ul>)(.*?)(</ul>)', re.I | re.S)
    match = section.search(html)
    if not match:
        raise RuntimeError("missing reference section")
    if 'class="source-link"' in match.group(2):
        return html
    updated = match.group(1) + match.group(2).rstrip() + "\n" + replacement + "\n" + match.group(3)
    return html[: match.start()] + updated + html[match.end() :]


def main() -> int:
    parser = argparse.ArgumentParser(description="Link school tutoring pages to official and representative sources.")
    parser.add_argument("xlsx", type=Path, help="2025 school statistics workbook")
    args = parser.parse_args()

    rows = json.loads(SCHOOL_MAP.read_text(encoding="utf-8"))
    groups: dict[tuple[str, str], list[str]] = defaultdict(list)
    for row in rows:
        groups[(row["official_school_name"], row["city"])].append(row["keyword"])

    homepages = read_homepages(args.xlsx, set(groups))
    missing_homepages = sorted(set(groups) - set(homepages))
    if missing_homepages:
        raise RuntimeError(f"homepage match failed for {len(missing_homepages)} schools: {missing_homepages[:10]}")

    homepage_rows = []
    changed = 0
    verified = 0
    for (official_name, city), slugs in sorted(groups.items()):
        general = [slug for slug in slugs if not slug.endswith(("수학과외", "영어과외"))]
        subjects = [slug for slug in slugs if slug.endswith(("수학과외", "영어과외"))]
        if len(general) != 1 or len(subjects) != 2:
            raise RuntimeError(f"unexpected page group for {official_name}: {slugs}")
        general_slug = general[0]
        homepage = homepages[(official_name, city)]
        homepage_rows.append({"official_school_name": official_name, "city": city, "homepage": homepage, "page": general_slug})

        for slug in [general_slug, *subjects]:
            path = OUTPUT / slug / "index.html"
            if not path.exists():
                raise RuntimeError(f"missing output page: {path}")
            html = path.read_text(encoding="utf-8")
            if slug == general_slug:
                link = source_link(homepage, f"{official_name} 공식 홈페이지", True)
                updated = update_reference_section(html, link, "학교 공식 홈페이지 주소", True)
                expected_href = homepage
            else:
                href = f"/{general_slug}/"
                link = source_link(href, f"{general_slug} 페이지로 이동", False)
                updated = update_reference_section(html, link, "학교 기본정보 확인", False)
                expected_href = href
            updated = updated.replace("/assets/css/style.css?v=school-link-preview-2", "/assets/css/style.css")
            if updated != html:
                path.write_text(updated, encoding="utf-8")
                changed += 1
            check = path.read_text(encoding="utf-8")
            if check.count('class="source-link"') != 1 or f'href="{expected_href}"' not in check:
                raise RuntimeError(f"link verification failed: {slug}")
            verified += 1

    HOMEPAGE_MAP.write_text(json.dumps(homepage_rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"schools": len(groups), "pages_verified": verified, "pages_changed": changed}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
