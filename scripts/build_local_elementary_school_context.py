from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = ROOT / "output"
DEFAULT_OUTPUT = ROOT / "data" / "local_elementary_school_context.json"
TARGET_PATTERN = re.compile(r"^(부산|양산|구미).+(?:동|읍|면)초등수학과외$")


def _integer(value: Any) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _target_locations() -> dict[str, tuple[str, str]]:
    targets: dict[str, tuple[str, str]] = {}
    for path in sorted(OUTPUT_ROOT.glob("*/index.html")):
        slug = path.parent.name
        match = TARGET_PATTERN.fullmatch(slug)
        if not match:
            continue
        city = match.group(1)
        location = slug.removesuffix("초등수학과외")
        targets[location] = (city, location.removeprefix(city))
    return targets


def _matches_city(city: str, province: str, district: str, address: str) -> bool:
    if city == "부산":
        return province == "부산" or "부산광역시" in address
    if city == "양산":
        return district == "양산시" or "양산시" in address
    if city == "구미":
        return district == "구미시" or "구미시" in address
    return False


def build_context(workbook_path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["학교별 주요통계"]
    headers = {
        str(sheet.cell(17, column).value).replace("\n", " ").strip(): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(17, column).value
    }

    def value(row: tuple[Any, ...], name: str) -> Any:
        return row[headers[name] - 1]

    targets = _target_locations()
    locations: dict[str, list[dict[str, object]]] = {location: [] for location in targets}
    for row in sheet.iter_rows(min_row=18, values_only=True):
        if str(value(row, "학교급") or "").strip() != "초등학교":
            continue
        if "폐" in str(value(row, "상태") or ""):
            continue
        province = str(value(row, "시도") or "").strip()
        district = str(value(row, "행정구") or "").strip()
        address = str(value(row, "주소") or "").strip()
        for location, (city, town) in targets.items():
            if not _matches_city(city, province, district, address):
                continue
            if town not in address:
                continue
            homepage = str(value(row, "홈페이지") or "").strip()
            if homepage and not homepage.startswith(("http://", "https://")):
                homepage = f"https://{homepage}"
            locations[location].append(
                {
                    "name": str(value(row, "학교명") or "").strip(),
                    "address": address,
                    "homepage": homepage,
                    "students": _integer(value(row, "학생수_총계_계")),
                    "classes": _integer(value(row, "편성학급수_계")),
                    "grade_students": {
                        str(grade): _integer(value(row, f"{grade}학년_학생수_계"))
                        for grade in range(1, 7)
                    },
                }
            )

    for schools in locations.values():
        schools.sort(key=lambda school: str(school["name"]))
    workbook.close()
    return {
        "source": {
            "file": workbook_path.name,
            "survey_date": "2025-04-01",
            "extracted_date": "2026-02-06",
            "publisher": "한국교육개발원 교육데이터연구본부 국가교육통계센터",
            "mapping_rule": "학교 주소에 대상 읍·면·동 명칭이 포함되는 초등학교",
        },
        "locations": locations,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    context = build_context(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(context, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    locations = context["locations"]
    assert isinstance(locations, dict)
    counts = [len(schools) for schools in locations.values()]
    print(f"locations={len(locations)}")
    print(f"locations_with_schools={sum(count > 0 for count in counts)}")
    print(f"schools={sum(counts)}")
    print(f"max_schools_per_location={max(counts, default=0)}")
    print(f"output={args.output}")
    return 0 if len(locations) == 69 else 1


if __name__ == "__main__":
    raise SystemExit(main())
