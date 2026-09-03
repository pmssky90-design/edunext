from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "output"
CONTEXT_PATH = ROOT / "data" / "local_middle_school_context.json"
TARGET = re.compile(r"^(부산|구미|양산)(.+(?:동|읍|면))중등영어과외$")
CITY_FILTERS = {
    "부산": ("부산", None),
    "구미": ("경북", "구미시"),
    "양산": ("경남", "양산시"),
}


def normalized_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def valid_homepage(value: object) -> str:
    url = str(value or "").strip()
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""
    if parsed.netloc == "school.busanedu.net" and parsed.path.rstrip("/") == "":
        return ""
    return url


def address_matches_town(address: str, town: str) -> bool:
    # 학교 주소의 행정동·읍·면 표기를 기준으로만 일치시킨다. 도로명 일부에
    # 같은 글자가 포함된 경우는 지역 학교로 오인하지 않는다.
    pattern = rf"(?:\(|\s){re.escape(town)}(?:\d+가)?(?:[.\s)]|$)"
    return bool(re.search(pattern, address))


def parent_slug(path: Path) -> str:
    html = path.read_text(encoding="utf-8")
    match = re.search(r'<nav class="breadcrumb".*?</nav>', html, flags=re.I | re.S)
    if not match:
        return ""
    parents = re.findall(r'href="/([^"/]+)/"', match.group(0), flags=re.I)
    return parents[-1] if parents else ""


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build verified middle-school context for local middle-English pages."
    )
    parser.add_argument("xlsx", type=Path, help="2025 school statistics workbook")
    args = parser.parse_args()

    targets: dict[str, dict[str, str]] = {}
    for path in sorted(OUTPUT.glob("*/index.html")):
        match = TARGET.fullmatch(path.parent.name)
        if not match:
            continue
        city, town = match.groups()
        targets[path.parent.name] = {
            "city": city,
            "town": town,
            "parent_slug": parent_slug(path),
        }
    if len(targets) != 69:
        raise RuntimeError(f"expected 69 target pages, found {len(targets)}")

    workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row = next(sheet.iter_rows(min_row=17, max_row=17, values_only=True))
    headers = {normalized_header(value): index for index, value in enumerate(header_row)}
    required = {"시도", "행정구", "학교급", "학교명", "주소", "홈페이지"}
    missing = required - set(headers)
    if missing:
        raise RuntimeError(f"missing workbook columns: {sorted(missing)}")

    school_rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=18, values_only=True):
        if str(row[headers["학교급"]] or "").strip() != "중학교":
            continue
        school_rows.append(
            {
                "province": str(row[headers["시도"]] or "").strip(),
                "district": str(row[headers["행정구"]] or "").strip(),
                "school_name": str(row[headers["학교명"]] or "").strip(),
                "address": str(row[headers["주소"]] or "").strip(),
                "homepage": valid_homepage(row[headers["홈페이지"]]),
            }
        )
    workbook.close()

    peers_by_parent: dict[str, list[str]] = {}
    for slug, target in targets.items():
        peers_by_parent.setdefault(target["parent_slug"], []).append(slug)
    for slugs in peers_by_parent.values():
        slugs.sort()

    context: dict[str, object] = {
        "source": {
            "workbook": args.xlsx.name,
            "sheet": sheet.title,
            "header_row": 17,
            "school_level": "중학교",
        },
        "pages": {},
    }
    matched_pages = 0
    matched_schools = 0
    for slug, target in targets.items():
        city = target["city"]
        town = target["town"]
        province, district = CITY_FILTERS[city]
        schools: list[dict[str, str]] = []
        seen: set[str] = set()
        for row in school_rows:
            if row["province"] != province:
                continue
            if district and row["district"] != district:
                continue
            if not address_matches_town(row["address"], town):
                continue
            school_name = row["school_name"]
            if not school_name or school_name in seen:
                continue
            seen.add(school_name)
            schools.append(
                {
                    "school_name": school_name,
                    "homepage": row["homepage"],
                }
            )
        schools.sort(key=lambda item: item["school_name"])
        if schools:
            matched_pages += 1
            matched_schools += len(schools)

        same_parent = peers_by_parent.get(target["parent_slug"], [])
        peer_slugs: list[str] = []
        if len(same_parent) > 1:
            index = same_parent.index(slug)
            candidates = [same_parent[(index - 1) % len(same_parent)], same_parent[(index + 1) % len(same_parent)]]
            peer_slugs = list(dict.fromkeys(item for item in candidates if item != slug))

        context["pages"][slug] = {
            "city": city,
            "town": town,
            "parent_slug": target["parent_slug"],
            "peer_slugs": peer_slugs,
            "schools": schools,
        }

    CONTEXT_PATH.write_text(
        json.dumps(context, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    print(
        json.dumps(
            {
                "pages": len(targets),
                "pages_with_exact_town_middle_schools": matched_pages,
                "exact_town_middle_schools": matched_schools,
                "output": str(CONTEXT_PATH),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
