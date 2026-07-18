from __future__ import annotations

import csv
import hashlib
import html
import os
import re
import shutil
import subprocess
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from statistics import mean
from xml.etree import ElementTree

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
AUDIT = ROOT / "audit"
SNAPSHOT = ROOT / "data" / "source_snapshot"
VERIFY = ROOT / "verification" / "attached-source-before-after"
PYTHON = Path(r"C:\Users\user\AppData\Local\Programs\Python\Python314\python.exe")

CONTENT_CANDIDATES = [
    Path(r"C:\gptwp\자료\부산_구미_양산 메인허브키워드 학교 포함.xlsx"),
    Path(r"C:\Users\user\Downloads\부산_구미_양산_고등영어과외_시트_B열_완성 (1).xlsx"),
]
REGION_CANDIDATES = [
    Path(r"C:\gptwp\자료\부산 구미 양산 포함 경산 (메인 키워드).xlsx"),
    Path(r"C:\gptwp\자료\부산 구미 양산 포항 경산 (메인 키워드).xlsx"),
    Path(r"C:\Users\user\Downloads\부산 구미 양산 포항 경산 (메인 키워드)(4).xlsx"),
]
CONTENT_SNAPSHOT = SNAPSHOT / "부산_구미_양산 메인허브키워드 학교 포함(1).xlsx"
REGION_SNAPSHOT = SNAPSHOT / "부산 구미 양산 포항 경산 (메인 키워드)(4).xlsx"
ALLOWED_TAGS = {"h2", "h3", "h4", "p", "ul", "ol", "li", "blockquote", "strong", "b", "em", "br", "table", "thead", "tbody", "tr", "th", "td"}
NON_CONTENT_SHEETS = {"주요고등학교 지역매핑"}


@dataclass
class SourceRow:
    keyword: str
    sheet: str
    row: int
    keyword_col: int
    content_col: int
    raw_content: str
    raw_text: str
    raw_sha256: str
    page_type: str


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def pick_existing(candidates: list[Path]) -> Path:
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("No candidate source file exists: " + ", ".join(str(item) for item in candidates))


def col_name(index: int) -> str:
    return openpyxl.utils.get_column_letter(index)


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def strip_text(value: str) -> str:
    value = re.sub(r"(?is)<script.*?</script>|<style.*?</style>|<iframe.*?</iframe>|<object.*?</object>|<embed.*?</embed>|<form.*?</form>", " ", value)
    value = re.sub(r"(?i)<br\s*/?>", "\n", value)
    value = re.sub(r"<[^>]+>", " ", value)
    return re.sub(r"\s+", " ", html.unescape(value)).strip()


def count_html(value: str, tag: str) -> int:
    return len(re.findall(rf"(?i)<{tag}\b", value))


def sanitize_html(value: str) -> str:
    value = re.sub(r"(?is)<(script|iframe|object|embed|form)\b.*?</\1>", " ", value)
    value = re.sub(r"(?is)<(script|iframe|object|embed|form)\b[^>]*>", " ", value)
    value = re.sub(r"(?i)\s+on[a-z]+\s*=\s*(['\"]).*?\1", "", value)
    value = re.sub(r"(?i)\s+href\s*=\s*(['\"])\s*javascript:.*?\1", "", value)
    value = re.sub(r"(?i)<\s*/?\s*h1\b", lambda m: m.group(0).lower().replace("h1", "h2"), value)

    def keep_or_unwrap(match: re.Match[str]) -> str:
        slash, tag, attrs = match.group(1), match.group(2).lower(), match.group(3) or ""
        if tag not in ALLOWED_TAGS:
            return ""
        if slash:
            return f"</{tag}>"
        safe_attrs = ""
        if tag in {"td", "th"}:
            colspan = re.search(r'(?i)\scolspan=["\']?([0-9]+)', attrs)
            rowspan = re.search(r'(?i)\srowspan=["\']?([0-9]+)', attrs)
            if colspan:
                safe_attrs += f' colspan="{colspan.group(1)}"'
            if rowspan:
                safe_attrs += f' rowspan="{rowspan.group(1)}"'
        return f"<{tag}{safe_attrs}>"

    value = re.sub(r"<\s*(/?)\s*([a-zA-Z0-9]+)\b([^>]*)>", keep_or_unwrap, value)
    if "<" not in value and ">" not in value:
        paragraphs = [item.strip() for item in re.split(r"\n{2,}", value) if item.strip()]
        return "".join(f"<p>{html.escape(item)}</p>" for item in paragraphs)
    return value.strip()


def detect_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int, int, int, int, float, int, int, bool, bool]:
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    column_scores = []
    formula = False
    for col in range(1, max_col + 1):
        non_empty = 0
        html_like = 0
        text_lengths = []
        for row in range(2, max_row + 1):
            cell = ws.cell(row, col)
            value = clean_cell(cell.value)
            if value:
                non_empty += 1
                text_lengths.append(len(strip_text(value)))
                if "<" in value and ">" in value:
                    html_like += 1
            if cell.data_type == "f":
                formula = True
        avg_len = mean(text_lengths) if text_lengths else 0
        column_scores.append((col, non_empty, avg_len, html_like))
    content_col = max(column_scores, key=lambda item: (item[2], item[3], item[1]))[0]
    keyword_candidates = [item for item in column_scores if item[0] != content_col]
    keyword_col = max(keyword_candidates, key=lambda item: (item[1], -item[2]))[0] if keyword_candidates else 1
    body_rows = 0
    empty_rows = 0
    lengths = []
    html_count = 0
    for row in range(2, max_row + 1):
        keyword = clean_cell(ws.cell(row, keyword_col).value)
        body = clean_cell(ws.cell(row, content_col).value)
        if not keyword:
            continue
        if body:
            body_rows += 1
            text = strip_text(body)
            lengths.append(len(text))
            if "<" in body and ">" in body:
                html_count += 1
        else:
            empty_rows += 1
    avg = mean(lengths) if lengths else 0
    return keyword_col, content_col, body_rows, empty_rows, int(avg), min(lengths or [0]), max(lengths or [0]), html_count > 0, formula


def workbook_evidence(path: Path, include_sheet_stats: bool) -> tuple[list[str], dict[str, tuple[int, int]]]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    lines = [
        f"- 실제 읽은 파일명: {path.name}",
        f"- 실제 경로: {path}",
        f"- 파일 크기: {path.stat().st_size}",
        f"- SHA-256: {sha256_file(path)}",
        f"- 시트 수: {len(wb.sheetnames)}",
        f"- 전체 시트명: {', '.join(wb.sheetnames)}",
    ]
    sheet_dims = {}
    for ws in wb.worksheets:
        sheet_dims[ws.title] = (ws.max_row or 0, ws.max_column or 0)
        lines.append(f"  - {ws.title}: max_row={ws.max_row}, max_col={ws.max_column}")
        if include_sheet_stats:
            k_col, c_col, body_rows, empty_rows, avg_len, min_len, max_len, has_html, has_formula = detect_columns(ws)
            lines.append(
                "    "
                f"keyword_col={col_name(k_col)}, content_col={col_name(c_col)}, "
                f"body_rows={body_rows}, empty_rows={empty_rows}, avg_text_len={avg_len}, "
                f"min_text_len={min_len}, max_text_len={max_len}, html={has_html}, formula={has_formula}"
            )
    return lines, sheet_dims


def page_type(keyword: str, sheet: str) -> str:
    if "학교" in sheet or keyword.endswith(("고과외", "고수학과외", "고영어과외")):
        return "school"
    if any(item in keyword for item in ["초등", "중등", "고등"]) and any(item in keyword for item in ["영어", "수학"]):
        return "subject_grade"
    if any(item in keyword for item in ["초등", "중등", "고등"]):
        return "grade"
    if any(item in keyword for item in ["영어", "수학"]):
        return "subject"
    return "region"


def load_source_rows(path: Path) -> tuple[list[SourceRow], list[dict[str, object]]]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    rows: list[SourceRow] = []
    sheet_stats = []
    for ws in wb.worksheets:
        k_col, c_col, body_rows, empty_rows, avg_len, min_len, max_len, has_html, has_formula = detect_columns(ws)
        sheet_stats.append({
            "sheet": ws.title,
            "keyword_col": col_name(k_col),
            "content_col": col_name(c_col),
            "body_rows": body_rows,
            "empty_rows": empty_rows,
            "avg_text_len": avg_len,
            "min_text_len": min_len,
            "max_text_len": max_len,
            "html": has_html,
            "formula": has_formula,
        })
        if ws.title in NON_CONTENT_SHEETS:
            continue
        for row_idx in range(2, (ws.max_row or 1) + 1):
            keyword = re.sub(r"\s+", "", clean_cell(ws.cell(row_idx, k_col).value))
            raw = clean_cell(ws.cell(row_idx, c_col).value)
            if not keyword:
                continue
            raw_text = strip_text(raw)
            rows.append(SourceRow(keyword, ws.title, row_idx, k_col, c_col, raw, raw_text, sha256_text(raw_text), page_type(keyword, ws.title)))
    return rows, sheet_stats


def extract_article_text(path: Path) -> tuple[str, str]:
    if not path.exists():
        return "", ""
    text = path.read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'(?is)<article class="content-body">(.*?)</article>', text)
    if not match:
        match = re.search(r"(?is)<main\b[^>]*>(.*?)</main>", text)
    article = match.group(1) if match else ""
    article = re.sub(r'(?is)<section class="link-section".*', "", article)
    return article, strip_text(article)


def snippet(text: str, where: str) -> str:
    if not text:
        return ""
    if where == "start":
        return text[:100]
    if where == "middle":
        start = max(0, len(text) // 2 - 50)
        return text[start:start + 100]
    return text[-100:]


def contains_near(haystack: str, needle: str) -> bool:
    needle = re.sub(r"\s+", " ", needle).strip()
    if not needle:
        return True
    return needle[:80] in haystack or needle[:50] in haystack


def fidelity_record(row: SourceRow, output_root: Path) -> dict[str, object]:
    path = output_root / row.keyword / "index.html"
    article_html, text = extract_article_text(path)
    raw_len = len(row.raw_text)
    page_len = len(text)
    ratio = round((page_len / raw_len * 100), 2) if raw_len else 0
    if ratio >= 90:
        verdict = "정상"
    elif ratio >= 70:
        verdict = "주의"
    elif ratio >= 40:
        verdict = "불량"
    else:
        verdict = "심각"
    return {
        "keyword": row.keyword,
        "source_sheet": row.sheet,
        "source_row": row.row,
        "source_column": col_name(row.content_col),
        "raw_character_count": raw_len,
        "page_character_count": page_len,
        "character_diff": page_len - raw_len,
        "preservation_ratio": ratio,
        "verdict": verdict,
        "raw_paragraph_count": count_html(row.raw_content, "p"),
        "page_paragraph_count": count_html(article_html, "p"),
        "raw_heading_count": len(re.findall(r"(?i)<h[1-6]\b", row.raw_content)),
        "raw_list_count": count_html(row.raw_content, "ul") + count_html(row.raw_content, "ol") + count_html(row.raw_content, "li"),
        "start_preserved": contains_near(text, snippet(row.raw_text, "start")),
        "middle_preserved": contains_near(text, snippet(row.raw_text, "middle")),
        "end_preserved": contains_near(text, snippet(row.raw_text, "end")),
        "fallback_used": bool(row.raw_text and "학부모 확인 기준" in text and row.raw_text[:40] not in text),
        "output_exists": path.exists(),
        "output_path": str(path),
        "raw_sha256": row.raw_sha256,
        "page_sha256": sha256_text(text),
    }


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not fields:
        fields = list(rows[0].keys()) if rows else ["empty"]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def copy_snapshot(src: Path, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)


def make_page_map(rows: list[SourceRow]) -> list[dict[str, object]]:
    result = []
    for row in rows:
        output = ROOT / "output" / row.keyword / "index.html"
        result.append({
            "excel_keyword": row.keyword,
            "source_sheet": row.sheet,
            "source_row": row.row,
            "content_cell": f"{col_name(row.content_col)}{row.row}",
            "output_path": str(output),
            "output_exists": output.exists(),
            "page_type": row.page_type,
            "is_region_page": row.page_type == "region",
            "is_school_page": row.page_type == "school",
        })
    return result


def extra_output_pages(rows: list[SourceRow]) -> list[dict[str, object]]:
    source_keywords = {row.keyword for row in rows}
    extras = []
    for path in (ROOT / "output").glob("*/index.html"):
        keyword = path.parent.name
        if keyword not in source_keywords:
            extras.append({"keyword": keyword, "output_path": str(path)})
    return extras


def pipeline_trace(rows: list[SourceRow], sample_limit: int = 60) -> list[dict[str, object]]:
    selected = []
    wanted_types = ["region", "subject", "grade", "subject_grade", "school"]
    for kind in wanted_types:
        selected.extend([row for row in rows if row.page_type == kind][:12])
    selected = selected[:sample_limit]
    records = []
    for row in selected:
        sanitized = sanitize_html(row.raw_content)
        rendered_html, rendered_text = extract_article_text(ROOT / "output_content_fixed" / row.keyword / "index.html")
        for stage, value in [
            ("excel_cell", row.raw_content),
            ("loader_raw_text", row.raw_text),
            ("sanitized_html", sanitized),
            ("rendered_article", rendered_text),
        ]:
            text = strip_text(value)
            records.append({
                "keyword": row.keyword,
                "page_type": row.page_type,
                "stage": stage,
                "length": len(text),
                "sha256": sha256_text(text),
                "start_100": snippet(text, "start"),
                "middle_100": snippet(text, "middle"),
                "end_100": snippet(text, "end"),
            })
    return records


def make_verification_samples(rows: list[SourceRow], before: list[dict[str, object]], after: list[dict[str, object]]) -> None:
    VERIFY.mkdir(parents=True, exist_ok=True)
    before_by = {item["keyword"]: item for item in before}
    after_by = {item["keyword"]: item for item in after}
    selected = []
    for kind in ["region", "subject", "grade", "subject_grade", "school"]:
        for city in ["부산", "구미", "양산"]:
            match = next((row for row in rows if row.page_type == kind and row.keyword.startswith(city)), None)
            if match and match.keyword not in {item.keyword for item in selected}:
                selected.append(match)
    lines = ["# Attached Source Before/After PC Comparison", ""]
    for row in selected[:24]:
        before_html, before_text = extract_article_text(ROOT / "output" / row.keyword / "index.html")
        after_html, after_text = extract_article_text(ROOT / "output_content_fixed" / row.keyword / "index.html")
        safe = re.sub(r"[^0-9A-Za-z가-힣_-]+", "_", row.keyword)
        (VERIFY / f"{safe}-before-article.html").write_text(before_html, encoding="utf-8")
        (VERIFY / f"{safe}-after-article.html").write_text(after_html, encoding="utf-8")
        lines.extend([
            f"## {row.keyword}",
            f"- source_sheet: {row.sheet}",
            f"- raw_chars: {len(row.raw_text)}",
            f"- before_chars: {len(before_text)}",
            f"- after_chars: {len(after_text)}",
            f"- before_ratio: {before_by.get(row.keyword, {}).get('preservation_ratio', '')}",
            f"- after_ratio: {after_by.get(row.keyword, {}).get('preservation_ratio', '')}",
            f"- raw_start: {snippet(row.raw_text, 'start')}",
            f"- raw_middle: {snippet(row.raw_text, 'middle')}",
            f"- raw_end: {snippet(row.raw_text, 'end')}",
            "",
        ])
    (VERIFY / "comparison-summary.md").write_text("\n".join(lines), encoding="utf-8")


def summarize_fidelity(rows: list[dict[str, object]]) -> dict[str, object]:
    with_content = [row for row in rows if int(row["raw_character_count"]) > 0]
    ratios = [float(row["preservation_ratio"]) for row in with_content]
    return {
        "source_pages": len(with_content),
        "avg_raw_chars": round(mean([int(row["raw_character_count"]) for row in with_content]), 2) if with_content else 0,
        "avg_page_chars": round(mean([int(row["page_character_count"]) for row in with_content]), 2) if with_content else 0,
        "avg_ratio": round(mean(ratios), 2) if ratios else 0,
        "under_90": sum(1 for row in with_content if float(row["preservation_ratio"]) < 90),
        "under_70": sum(1 for row in with_content if float(row["preservation_ratio"]) < 70),
        "under_40": sum(1 for row in with_content if float(row["preservation_ratio"]) < 40),
        "fallback_used": sum(1 for row in with_content if str(row["fallback_used"]).lower() == "true"),
    }


def update_source_content(rows: list[SourceRow]) -> None:
    # The generator already stores full source body in Page.body. Sanitization is handled in data_loader/render path.
    pass


def run_generator_content_fixed() -> None:
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["EDUNEXT_OUTPUT"] = "content_fixed"
    env["EDUNEXT_STRICT_SOURCE"] = "1"
    subprocess.run([str(PYTHON), "generator.py"], cwd=ROOT, env=env, check=True)


def main() -> int:
    AUDIT.mkdir(exist_ok=True)
    SNAPSHOT.mkdir(parents=True, exist_ok=True)
    if os.environ.get("EDUNEXT_AFTER_ONLY") == "1":
        rows, _ = load_source_rows(pick_existing(CONTENT_CANDIDATES))
        source_rows = [row for row in rows if row.raw_content]
        after = [fidelity_record(row, ROOT / "output_content_fixed") for row in source_rows]
        write_csv(AUDIT / "attached-content-fidelity-after.csv", after)
        write_csv(AUDIT / "content-fidelity-after.csv", after)
        after_summary = summarize_fidelity(after)
        print("after:", after_summary)
        return 0
    content_src = pick_existing(CONTENT_CANDIDATES)
    region_src = pick_existing(REGION_CANDIDATES)
    copy_snapshot(content_src, CONTENT_SNAPSHOT)
    copy_snapshot(region_src, REGION_SNAPSHOT)

    evidence = ["# Source Snapshot", "", "## Content Workbook"]
    content_lines, _ = workbook_evidence(content_src, True)
    region_lines, _ = workbook_evidence(region_src, False)
    evidence.extend(content_lines)
    evidence.extend(["", "## Region Workbook"])
    evidence.extend(region_lines)
    evidence.extend([
        "",
        "## Snapshot Integrity",
        f"- content_original_sha256: {sha256_file(content_src)}",
        f"- content_snapshot_sha256: {sha256_file(CONTENT_SNAPSHOT)}",
        f"- region_original_sha256: {sha256_file(region_src)}",
        f"- region_snapshot_sha256: {sha256_file(REGION_SNAPSHOT)}",
        f"- content_original_path: {content_src}",
        f"- region_original_path: {region_src}",
    ])
    (AUDIT / "source-snapshot.md").write_text("\n".join(evidence) + "\n", encoding="utf-8")

    rows, sheet_stats = load_source_rows(content_src)
    source_rows = [row for row in rows if row.raw_content]
    write_csv(AUDIT / "attached-sheet-column-analysis.csv", sheet_stats)
    write_csv(AUDIT / "attached-source-page-map.csv", make_page_map(rows))
    write_csv(AUDIT / "attached-extra-output-pages.csv", extra_output_pages(rows))

    before = [fidelity_record(row, ROOT / "output") for row in source_rows]
    write_csv(AUDIT / "attached-content-fidelity-before.csv", before)

    missing = [
        {"keyword": row.keyword, "source_sheet": row.sheet, "source_row": row.row}
        for row in rows
        if not row.raw_content
    ]
    too_short = [
        {
            "keyword": row.keyword,
            "source_sheet": row.sheet,
            "source_row": row.row,
            "raw_character_count": len(row.raw_text),
            "paragraph_count": count_html(row.raw_content, "p"),
        }
        for row in rows
        if row.raw_content and (len(row.raw_text) < 500 or count_html(row.raw_content, "p") < 3)
    ]
    write_csv(AUDIT / "attached-source-content-missing.csv", missing, ["keyword", "source_sheet", "source_row"])
    write_csv(AUDIT / "attached-source-content-too-short.csv", too_short)

    if os.environ.get("EDUNEXT_SKIP_GENERATE") != "1":
        run_generator_content_fixed()
    after = [fidelity_record(row, ROOT / "output_content_fixed") for row in source_rows]
    write_csv(AUDIT / "attached-content-fidelity-after.csv", after)
    write_csv(AUDIT / "content-fidelity-after.csv", after)
    if os.environ.get("EDUNEXT_FAST") != "1":
        write_csv(AUDIT / "attached-content-pipeline-trace.csv", pipeline_trace(source_rows))
        make_verification_samples(source_rows, before, after)

    before_summary = summarize_fidelity(before)
    after_summary = summarize_fidelity(after)
    summary_lines = ["# Attached Content Recovery Summary", "", "## Before"]
    summary_lines.extend(f"- {key}: {value}" for key, value in before_summary.items())
    summary_lines.extend(["", "## After"])
    summary_lines.extend(f"- {key}: {value}" for key, value in after_summary.items())
    summary_lines.extend([
        "",
        "## Cause",
        "- 기존 생성기는 원문 본문을 article에 넣고 있었으나 입력 경로가 C:\\gptwp\\자료 고정이라 이번 첨부 원본 우선 사용 조건을 재현할 수 없었다.",
        "- 이번 수정으로 원본을 data/source_snapshot에 보관하고 config.py 한 곳에서 스냅샷을 기본 입력으로 사용한다.",
        "- article 추출 기준과 전체 본문 보존율 검증 CSV를 추가했다.",
    ])
    (AUDIT / "attached-content-recovery-summary.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")
    print("content_source:", content_src)
    print("region_source:", region_src)
    print("before:", before_summary)
    print("after:", after_summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
