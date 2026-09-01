from __future__ import annotations

from collections import OrderedDict
import re

import openpyxl

from config import CITY_PROVINCE, CONTENT_EXCEL, REGION_EXCEL, SCHOOL_SHEETS, TARGET_CITIES
from sitegen.models import Region
from sitegen.utils import clean, normalize_slug


SOURCE_TEXT_REPLACEMENTS = {
    "학습 메모하는가": "기록하는가",
    "학습 메모한다": "기록한다",
    "학습 메모하고": "기록하고",
    "학습 메모에서": "기록에서",
    "학습 메모을": "기록을",
    "학습 메모과": "기록과",
    "학습 메모은": "기록은",
    "학습 메모해": "기록해",
    "시험 학습가": "시험 학습이",
    "영어 학습는": "영어 학습은",
    "출발점에서는 읽고": "출발점으로 읽고",
    "처음에는 정한다": "먼저 정한다",
    "영어학습": "영어 학습",
    "수학학습": "수학 학습",
}


def normalize_source_text(body: str) -> str:
    """Repair known malformed copy before imported HTML becomes page content."""
    for malformed, corrected in SOURCE_TEXT_REPLACEMENTS.items():
        body = body.replace(malformed, corrected)
    return re.sub(
        r"(초등학생|중학생|고등학생)에게는\s+\1은",
        r"\1은",
        body,
    )


def sanitize_source_html(body: str) -> str:
    body = re.sub(r"(?is)<(script|iframe|object|embed|form)\b.*?</\1>", " ", body)
    body = re.sub(r"(?is)<(script|iframe|object|embed|form)\b[^>]*>", " ", body)
    body = re.sub(r"(?i)\s+on[a-z]+\s*=\s*(['\"]).*?\1", "", body)
    body = re.sub(r"(?i)\s+href\s*=\s*(['\"])\s*javascript:.*?\1", "", body)
    body = re.sub(r"(?i)<\s*/?\s*h1\b", lambda m: m.group(0).lower().replace("h1", "h2"), body)
    body = re.sub(r'<a\b([^>]*?)\s+href=["\']https?://[^"\']+["\']([^>]*)>', r"<span\1\2>", body, flags=re.I)
    body = re.sub(r"</a>", "</span>", body, flags=re.I)
    return normalize_source_text(body)


def add_region(regions: OrderedDict[str, Region], key: str, name: str, slug: str, level: str, parent: str | None) -> None:
    if key not in regions:
        regions[key] = Region(key=key, name=name, slug=slug, level=level, parent=parent)
    if parent and key not in regions[parent].children:
        regions[parent].children.append(key)


def load_regions() -> OrderedDict[str, Region]:
    regions: OrderedDict[str, Region] = OrderedDict()
    add_region(regions, "national", "전국", "전국", "national", None)
    for province in ["부산", "경남", "경북"]:
        add_region(regions, f"province:{province}", province, province, "province", "national")

    wb = openpyxl.load_workbook(REGION_EXCEL, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        city = clean(row[0] if len(row) > 0 else "")
        district = clean(row[1] if len(row) > 1 else "")
        town = clean(row[2] if len(row) > 2 else "")
        if city not in TARGET_CITIES or not town:
            continue
        province = CITY_PROVINCE[city]
        if city == "부산":
            city_key = "province:부산"
        else:
            city_key = f"city:{city}"
            add_region(regions, city_key, city, city, "city", f"province:{province}")
        parent_key = city_key
        if district:
            district_key = f"district:{city}:{district}"
            district_slug = district if district.startswith(city) else f"{city}{district}"
            add_region(regions, district_key, district, district_slug, "district", city_key)
            parent_key = district_key
        town_key = f"town:{city}:{district}:{town}"
        town_slug = town if town.startswith(city) else f"{city}{town}"
        add_region(regions, town_key, town, town_slug, "town", parent_key)
    return regions


def load_content() -> tuple[dict[str, str], set[str]]:
    content: dict[str, str] = {}
    school_slugs: set[str] = set()
    sources: dict[str, str] = {}
    non_content_sheets = {"주요고등학교 지역매핑"}
    wb = openpyxl.load_workbook(CONTENT_EXCEL, read_only=True, data_only=True)
    for ws in wb.worksheets:
        if ws.title in non_content_sheets:
            continue
        is_school = ws.title in SCHOOL_SHEETS
        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index == 0:
                continue
            keyword = normalize_slug(clean(row[0] if row else ""))
            body = clean(row[1] if len(row) > 1 else "")
            if not keyword or not body:
                continue
            if is_school and not keyword.startswith(tuple(TARGET_CITIES)):
                continue
            content[keyword] = sanitize_source_html(body)
            sources[keyword] = ws.title
            if is_school:
                school_slugs.add(keyword)
    load_content.sources = sources
    return content, school_slugs


load_content.sources = {}


def load_content_sources() -> dict[str, str]:
    if not load_content.sources:
        load_content()
    return dict(load_content.sources)


def load_school_region_map() -> dict[str, dict[str, str]]:
    wb = openpyxl.load_workbook(CONTENT_EXCEL, read_only=True, data_only=True)
    if "주요고등학교 지역매핑" not in wb.sheetnames:
        return {}
    ws = wb["주요고등학교 지역매핑"]
    mapping: dict[str, dict[str, str]] = {}
    for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if index == 1:
            continue
        city = clean(row[4] if len(row) > 4 else "")
        district = clean(row[5] if len(row) > 5 else "")
        town = clean(row[6] if len(row) > 6 else "")
        display = clean(row[7] if len(row) > 7 else "")
        official = clean(row[8] if len(row) > 8 else "")
        if city not in TARGET_CITIES or not display:
            continue
        base = display if display.startswith(city) else f"{city}{display}"
        district_slug = district if district.startswith(city) else f"{city}{district}" if district else ""
        town_slug = town if town.startswith(city) else f"{city}{town}" if town else ""
        for suffix, kind in [("과외", "school_tutoring"), ("수학과외", "school_subject_math"), ("영어과외", "school_subject_english")]:
            keyword = f"{base}{suffix}"
            mapping[keyword] = {
                "keyword": keyword,
                "base": base,
                "page_type": kind,
                "school_display_name": display,
                "official_school_name": official,
                "city": city,
                "district": district,
                "town": town,
                "city_slug": city,
                "district_slug": district_slug,
                "town_slug": town_slug,
                "source_row": str(index),
            }
    return mapping
